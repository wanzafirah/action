"""Load TalentCorp staff list from the bundled Excel file."""
from __future__ import annotations

from functools import lru_cache
from pathlib import Path


_EXCEL_PATH = Path(__file__).parent.parent / "TC Staff Details.xlsx"


@lru_cache(maxsize=1)
def load_tc_staff() -> list[dict]:
    """Return [{name, email, tcid}] from TC Staff Details.xlsx."""
    if not _EXCEL_PATH.exists():
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(_EXCEL_PATH, read_only=True, data_only=True)
        # Find the sheet with Employee Name header
        ws = None
        for sheet in wb.worksheets:
            first_row = [c.value for c in next(sheet.iter_rows(max_row=1))]
            if any("Employee Name" in str(v) for v in first_row if v):
                ws = sheet
                break
        if ws is None:
            ws = wb.worksheets[0]

        staff: list[dict] = []
        header_done = False
        for row in ws.iter_rows(values_only=True):
            if not header_done:
                header_done = True
                continue
            name  = row[0] if len(row) > 0 else None
            email = row[1] if len(row) > 1 else None
            tcid  = row[2] if len(row) > 2 else None
            if name:
                staff.append({
                    "name":  str(name).strip(),
                    "email": str(email).strip() if email else "",
                    "tcid":  str(tcid).strip()  if tcid  else "",
                })
        return sorted(staff, key=lambda x: x["name"])
    except Exception:
        return []


def get_tc_names() -> list[str]:
    """Return just the sorted name strings for use in multiselect / selectbox."""
    return [s["name"] for s in load_tc_staff()]
